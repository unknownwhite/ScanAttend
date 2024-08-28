from flask import Flask, render_template, request, jsonify
import pymongo
import cv2
from pyzbar.pyzbar import decode
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
from datetime import datetime 

app = Flask(__name__)

client = pymongo.MongoClient("mongodb://localhost:27017/")
db = client['Employees']
collection = db['SST Students']

EXCEL_FILE = 'student_data.xlsx'

def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        headers = ['ERP ID', 'Name', 'Phone', 'Email', 'Branch', 'Timestamp']  
        for col_num, header in enumerate(headers, 1):
            sheet[f'{get_column_letter(col_num)}1'] = header
        workbook.save(EXCEL_FILE)

def save_to_excel(student):
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    timestamp = datetime.now().strftime('%d-%m-%y %H:%M:%S')  # Get the current time and date
    row = (student['erpid'], student['name'], student['phone'], student['email'], student['branch'], timestamp)  # Add timestamp to the row
    sheet.append(row)
    workbook.save(EXCEL_FILE)


def capture_barcode():
    global scanning
    cap = cv2.VideoCapture(0) 
    barcode_data = None
    while scanning:
        ret, frame = cap.read()
        if not ret:
            break
        barcodes = decode(frame)
        for barcode in barcodes:
            barcode_data = barcode.data.decode('utf-8')
            break
        if barcode_data:
            break
        cv2.imshow('Barcode Scanner', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    cap.release()
    cv2.destroyAllWindows()
    return barcode_data

@app.route('/scan', methods=['POST'])
def scan():
    global scanning
    scanning = True
    barcode_data = capture_barcode()
    scanning = False
    if barcode_data:
        student = collection.find_one({'erpid': int(barcode_data)}, {'_id': 0, 'erpid': 1, 'name': 1, 'phone': 1, 'email': 1, 'branch': 1})
        if student:
            return jsonify(student)
        else:
            return jsonify({'error': 'No student found with ERP ID: ' + barcode_data})
    else:
        return jsonify({'error': 'No barcode detected or scanning stopped'})

@app.route('/stop_scan', methods=['POST'])
def stop_scan():
    global scanning
    scanning = False
    return jsonify({'message': 'Scanning stopped'})

@app.route('/search', methods=['POST'])
def search():
    data = request.get_json()  # Changed from request.form.get() to request.get_json()
    erpid = data.get('erpid')  # Access the 'erpid' from the JSON data
    if erpid:
        student = collection.find_one({'erpid': int(erpid)}, {'_id': 0, 'erpid': 1, 'name': 1, 'phone': 1, 'email': 1, 'branch': 1})
        if student:
            return jsonify(student)
        else:
            return jsonify({'error': f'No student found with ERP ID: {erpid}'})
    return jsonify({'error': 'No ERP ID provided'})

@app.route('/save', methods=['POST'])
def save():
    student = request.json
    if student:
        save_to_excel(student)
        return jsonify({'success': 'Student data saved to Excel'})
    return jsonify({'error': 'No student data to save'})

@app.route('/', methods=['GET', 'POST'])
def display_students():
    return render_template('index.html')

if __name__ == "__main__":
    app.run(debug=True, port=5000)
